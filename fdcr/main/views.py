from django.shortcuts import render
from .models import UIManager,ReportTypes,DataManager
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_protect
import openpyxl
from django.shortcuts import render,HttpResponse,redirect


# Create your views here.

def index(request):
    title_text = UIManager.objects.get(UI_position=1)
    text2 = UIManager.objects.get(UI_position=2)
    report_type = ReportTypes.objects.all()
    context = {
        "title_text":title_text,
        "text2":text2,
        "report_type":report_type
        }            
    upload_excel("report.xlsx")
    return render(request,'index.html',context)

def upload_excel(file_path):
    if file_path.name.endswith('.xls') or file_path.name.endswith('.xlsx'):
        dataframe = openpyxl.load_workbook(file_path)
        dataframe1 = dataframe.active  
        heading = False
        status = "success"
        for i in range(1, dataframe1.max_row):
            number = dataframe1.cell(row = i, column = 1 ).value
            type = dataframe1.cell(row = i, column = 2 ).value

            if heading == True:
                print(number,"::::",type)
                DataManager.objects.create(report_number=number,is_active=True,report_type=type)
            else:
                if number == "Number" and type=="Type":
                    heading = True
                else :
                    status = "error"

def perform_import_view(request):
    if request.method == 'POST' and request.FILES['excel_file']:
        excel_file = request.FILES['excel_file']
        upload_excel(excel_file)
        

        return HttpResponse("Excel uploaded successfully")
    return render(request, 'import_form.html')

@csrf_protect
def verify(request):
    if request.method == 'POST':
        q = request.POST['q']
        report_type = request.POST['report_type']
        try:
            data = DataManager.objects.get(report_number=q,is_active=True,report_type=report_type)
            return JsonResponse({'status': 'success', 'message': 'Successfully Verified'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': 'Not Verified'})
